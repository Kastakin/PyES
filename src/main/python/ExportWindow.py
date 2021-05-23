from datetime import datetime

from openpyxl.utils import get_column_letter
from pandas import ExcelWriter
from PyQt5.QtWidgets import QFileDialog, QWidget

from ui.sssc_dataExport import Ui_ExportWindow
from utils_func import getColWidths


class ExportWindow(QWidget, Ui_ExportWindow):
    def __init__(self, parent):
        super().__init__()
        self.setupUi(self)

        self.export_path = None
        self.result = parent.result

        if "formation_constants" not in self.result:
            self.adjlogb_check.setChecked(False)
            self.adjlogb_check.setEnabled(False)

    def ExcelExport(self):
        if self.export_path:
            self.export_path, _ = QFileDialog.getSaveFileName(
                self, "Pick a Path", self.export_path, "Excel (*.xlsx)"
            )
        else:
            self.export_path, _ = QFileDialog.getSaveFileName(
                self, "Pick a Path", "", "Excel (*.xlsx)"
            )

        if self.export_path:
            self.export_path = self.export_path.split(".")[0] + ".xlsx"

            with ExcelWriter(self.export_path) as writer:
                wb = writer.book

                if self.input_check.isChecked():
                    self.result["species_info"].to_excel(
                        writer, sheet_name="System Info", startrow=3
                    )
                    self.result["comp_info"].to_excel(
                        writer,
                        sheet_name="System Info",
                        startrow=3,
                        startcol=(self.result["species_info"].shape[1] + 3),
                    )

                    ws = wb["System Info"]
                    ws["A1"] = "File:"
                    ws["A2"] = "Date:"
                    ws.merge_cells("B1:D1")
                    ws.merge_cells("B2:D2")
                    # TODO: make it print sensible info
                    ws["B1"] = "SAMPLE FILE PLACEHOLDER"
                    ws["B2"] = datetime.now()

                if self.distribution_check.isChecked():
                    self.result["distribution"].to_excel(
                        writer, sheet_name="Species Distribution"
                    )

                    ws = wb["Species Distribution"]
                    dist_widths = getColWidths(self.result["distribution"])
                    for i, column_width in enumerate(dist_widths):
                        ws.column_dimensions[
                            get_column_letter(i + 1)
                        ].width = column_width

                if self.adjlogb_check.isChecked():
                    self.result["formation_constants"].to_excel(
                        writer,
                        sheet_name="Adjusted Formation Constants",
                        float_format="%.3f",
                    )
                    ws = wb["Adjusted Formation Constants"]
                    logb_width = getColWidths(self.result["formation_constants"])
                    for i, column_width in enumerate(logb_width):
                        ws.column_dimensions[
                            get_column_letter(i + 1)
                        ].width = column_width

    def TableEnabler(self, state):
        if state == True:
            self.perc_group.setEnabled(True)
        else:
            self.perc_group.setEnabled(False)
